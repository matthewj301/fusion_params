#!/usr/bin/env python3
"""Convert FDM_MasterParams xlsx to validated CSV for Fusion 360 import.

Reads the xlsx (single source of truth), validates every row, topologically
sorts derived parameters after their dependencies, and writes a clean CSV.

Usage:
    python3 convert_xlsx.py [--xlsx PATH] [--output PATH]

Defaults:
    --xlsx   FDM_MasterParams_BakedBean3D_v4.xlsx
    --output BakedBean3D_MasterParams_v4/BakedBean3D_MasterParams_params.csv
"""

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


FUSION_NATIVE_UNITS = {"mm", "cm", "in", "deg", "rad", ""}
DIMENSIONLESS_UNITS = {"count", "x", "%", "°C", "A", "µm/mm"}
VALID_UNITS = FUSION_NATIVE_UNITS | DIMENSIONLESS_UNITS
UNIT_REMAP = {u: "" for u in DIMENSIONLESS_UNITS}

FUSION_UNIT_SUFFIXES = {"mm", "cm", "in", "deg", "rad"}

PARAM_NAME_RE = re.compile(r"^[a-z][a-z0-9_]*$")
HEADER_ROW_MARKER = "Parameter Name"


def find_data_start(ws):
    """Find the first row after the header row containing column labels."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] and HEADER_ROW_MARKER in str(row[0]):
            return i + 1
    return None


def clean_name(raw):
    """Strip emoji prefixes (⚠, etc.) and whitespace from parameter names."""
    return re.sub(r"^[⚠️\s]+", "", str(raw)).strip()


def extract_dependencies(expression):
    """Extract parameter name references from a Fusion expression.

    Derived expressions start with '=' in the xlsx and reference other
    param names like '2*printer_accuracy_extrusion_line_width'.
    Filters out unit suffixes (mm, deg, etc.) that aren't param names.
    """
    expr = str(expression).lstrip("=").strip()
    tokens = set(re.findall(r"[a-z][a-z0-9_]{2,}", expr))
    return tokens - FUSION_UNIT_SUFFIXES


def is_category_header(row):
    """Category headers have an emoji name in col A but no value/unit."""
    return row[1] is None and row[2] is None


def validate_and_parse(ws, data_start):
    """Parse and validate all parameter rows from the worksheet."""
    params = []
    errors = []
    seen_names = {}

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True),
        start=data_start,
    ):
        raw_name = row[0]
        if raw_name is None or str(raw_name).strip() == "":
            continue
        if is_category_header(row):
            continue

        name = clean_name(raw_name)
        value = row[1]
        unit = str(row[2]).strip() if row[2] else ""
        comment = str(row[3]).strip() if row[3] else ""
        expression = str(row[4]).strip() if row[4] else ""

        row_errors = []

        if not PARAM_NAME_RE.match(name):
            row_errors.append(f"invalid parameter name '{name}' (must be lowercase, underscores, start with letter)")

        if name in seen_names:
            row_errors.append(f"duplicate name '{name}' (first seen at row {seen_names[name]})")
        seen_names[name] = row_idx

        if value is None and not expression:
            row_errors.append("missing both Value and Expression")

        if unit not in VALID_UNITS:
            row_errors.append(f"invalid unit '{unit}' (allowed: {', '.join(sorted(VALID_UNITS)) or '(empty)'})")

        unit = UNIT_REMAP.get(unit, unit)

        if not expression:
            row_errors.append("missing Fusion expression (column E)")

        for e in row_errors:
            errors.append(f"  Row {row_idx}: {e}")

        expr_clean = expression.lstrip("=").strip()
        deps = extract_dependencies(expression)

        params.append({
            "name": name,
            "expression": expr_clean,
            "unit": unit,
            "comment": comment,
            "deps": deps,
            "row": row_idx,
        })

    return params, errors


def validate_references(params):
    """Check that all derived param dependencies reference existing params."""
    all_names = {p["name"] for p in params}
    errors = []
    for p in params:
        missing = p["deps"] - all_names
        if missing:
            errors.append(f"  Row {p['row']}: '{p['name']}' references undefined params: {', '.join(sorted(missing))}")
    return errors


def topological_sort(params):
    """Sort so that base params come before derived params that reference them.

    Detects cycles and raises if found.
    """
    name_to_param = {p["name"]: p for p in params}
    visited = set()
    in_stack = set()
    order = []

    def visit(name):
        if name in in_stack:
            raise ValueError(f"circular dependency involving '{name}'")
        if name in visited:
            return
        in_stack.add(name)
        p = name_to_param.get(name)
        if p:
            for dep in p["deps"]:
                if dep in name_to_param:
                    visit(dep)
        in_stack.remove(name)
        visited.add(name)
        order.append(name)

    for p in params:
        visit(p["name"])

    name_to_index = {name: i for i, name in enumerate(order)}
    return sorted(params, key=lambda p: name_to_index[p["name"]])


def write_csv(params, output_path):
    """Write validated, sorted params to CSV."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Expression", "Unit", "Comment"])
        for p in params:
            writer.writerow([p["name"], p["expression"], p["unit"], p["comment"]])


def main():
    parser = argparse.ArgumentParser(description="Convert FDM master params xlsx to validated CSV")
    parser.add_argument("--xlsx", default="FDM_MasterParams_BakedBean3D_v4.xlsx")
    parser.add_argument("--output", default="BakedBean3D_MasterParams_v4/BakedBean3D_MasterParams_params.csv")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
    ws = wb[wb.sheetnames[0]]

    data_start = find_data_start(ws)
    if not data_start:
        print("ERROR: Could not find header row with 'Parameter Name'", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {xlsx_path} (data starts at row {data_start})...")
    params, parse_errors = validate_and_parse(ws, data_start)
    ref_errors = validate_references(params)
    wb.close()

    all_errors = parse_errors + ref_errors
    if all_errors:
        print(f"\nFAILED — {len(all_errors)} validation error(s):", file=sys.stderr)
        for e in all_errors:
            print(e, file=sys.stderr)
        sys.exit(1)

    try:
        sorted_params = topological_sort(params)
    except ValueError as e:
        print(f"\nFAILED — {e}", file=sys.stderr)
        sys.exit(1)

    base_count = sum(1 for p in sorted_params if not p["deps"])
    derived_count = len(sorted_params) - base_count

    write_csv(sorted_params, output_path)
    print(f"Wrote {len(sorted_params)} parameters to {output_path}")
    print(f"  Base params: {base_count}")
    print(f"  Derived params: {derived_count}")


if __name__ == "__main__":
    main()
